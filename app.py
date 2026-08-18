import streamlit as st
import pandas as pd
import numpy as np
import requests
import json as _json
import base64
from openpyxl import load_workbook, Workbook
import io
import zipfile
from pathlib import Path
from datetime import datetime

st.set_page_config(page_title="WV BBL Tax Reporter", page_icon="🍺", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Libre+Baskerville:wght@400;700&family=Source+Code+Pro:wght@400;600&display=swap');
html, body, [class*="css"] { font-family: 'Libre Baskerville', serif; }
.stApp { background-color: #1a1208; color: #f0e6c8; }
h1 { font-family: 'Libre Baskerville', serif; color: #d4a843; font-size: 2rem;
     letter-spacing: 0.04em; border-bottom: 1px solid #4a3510; padding-bottom: 0.5rem; }
h2, h3 { font-family: 'Libre Baskerville', serif; color: #d4a843; }
.stButton > button { background-color: #d4a843; color: #1a1208; font-family: 'Libre Baskerville', serif;
     font-weight: 700; border: none; border-radius: 2px; padding: 0.5rem 1.5rem; letter-spacing: 0.05em; }
.stButton > button:hover { background-color: #e8c060; color: #1a1208; }
.stDownloadButton > button { background-color: #2a4a1a; color: #a0d080; font-family: 'Source Code Pro', monospace;
     border: 1px solid #4a7a2a; border-radius: 2px; width: 100%; }
.stDownloadButton > button:hover { background-color: #3a6a2a; }
.result-box { background: #241a08; border-left: 3px solid #d4a843; padding: 0.75rem 1rem; margin: 0.5rem 0;
     font-family: 'Source Code Pro', monospace; font-size: 0.85rem; color: #c8b888; }
.error-box { background: #2a0808; border-left: 3px solid #d44343; padding: 0.75rem 1rem; margin: 0.5rem 0;
     font-family: 'Source Code Pro', monospace; font-size: 0.85rem; color: #d44343; }
.template-saved { background: #0a2a1a; border-left: 3px solid #4a8a5a; padding: 0.5rem 1rem; margin: 0.5rem 0;
     font-family: 'Source Code Pro', monospace; font-size: 0.85rem; color: #80c880; }
.stExpander { border: 1px solid #4a3510 !important; background: #241a08 !important; }
label, .stFileUploader label { color: #c8b888 !important; }
.stDataFrame { font-family: 'Source Code Pro', monospace; font-size: 0.8rem; }
hr { border-color: #4a3510; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
GAL_PER_BBL = 31
FACILITY_CAPACITY_BBL = 5000
BREWER_NAME = "Mountain State Brewing Co."
LICENSE_NUMBER = "47-W-002-000225"          # built-in default if the sheet is unreachable
@st.cache_resource(show_spinner=False)
def _locate_template():
    """Find the upload template even if its name/location is slightly off."""
    here = Path(__file__).resolve().parent
    roots, checked = [here, here.parent, Path.cwd().resolve()], set()
    for root in roots:                       # exact name first
        if root in checked or not root.exists():
            continue
        checked.add(root)
        hits = list(root.rglob("__WV_Upload_Template.xlsx"))
        if hits:
            return str(hits[0])
    for root in roots:                       # fuzzy: any xlsx that looks like the template
        if not root.exists():
            continue
        for p in root.rglob("*.xlsx"):
            key = p.name.lower().replace("_", "").replace(" ", "")
            if "uploadtemplate" in key:
                return str(p)
    return None

_tpl = _locate_template()
TEMPLATE_PATH = Path(_tpl) if _tpl else None

# Template baked into the app (text module, commits reliably; no Drive/web-app needed).
try:
    from template_b64 import TEMPLATE_B64
    EMBEDDED_TEMPLATE = base64.b64decode(TEMPLATE_B64)
except Exception:
    EMBEDDED_TEMPLATE = None
HIST_HEADERS = ["Fiscal Year", "Total Sales (gallons)", "Total Sales (barrels)",
                "WV Distributor (gallons)", "Self-distributed (gallons)",
                "Brewpub (gallons)", "Other State (gallons)", "Generated"]

DEFAULT_ABCA = {
    "jefferson distributing co., inc.": ("02-T-001-000021", "Jefferson Distributing Co., Inc."),
    "northern eagle distributing":       ("14-T-001-000081", "Northern Eagle Distributing"),
    "beverage distributors":             ("17-T-001-000004", "Beverage Distributors"),
    "mountain eagle, inc.":              ("41-T-001-000028", "Mountain Eagle, Inc."),
    "carenbauer distributing":           ("35-T-001-000010", "Carenbauer Distributing"),
    "mona distributing":                 ("31-T-001-000027", "Mona Distributing"),
    "spriggs":                           ("54-T-001-017330", "Spriggs"),
}

if "abca_map" not in st.session_state:
    st.session_state.abca_map = DEFAULT_ABCA.copy()
if "results" not in st.session_state:
    st.session_state.results = []
if "eop" not in st.session_state:
    st.session_state.eop = None

# ── Storage layer: Google Apps Script web app ───────────────────────────────────
def _gas_cfg():
    try:
        return st.secrets["gas"]["url"], st.secrets["gas"]["token"]
    except Exception:
        return None, None

def storage_ready():
    url, token = _gas_cfg()
    return bool(url and token)

def _jsonable(o):
    if isinstance(o, np.integer):  return int(o)
    if isinstance(o, np.floating): return float(o)
    return str(o)

def _gas_post(payload):
    url, token = _gas_cfg()
    if not url or not token:
        return None
    try:
        body = _json.dumps({**payload, "token": token}, default=_jsonable)
        r = requests.post(url, data=body, headers={"Content-Type": "application/json"}, timeout=20)
        data = r.json()
        return data if data.get("ok") else None
    except Exception:
        return None

def gas_get_license():
    d = _gas_post({"action": "get_license"})
    return ((d or {}).get("license") or None)

def gas_set_license(value):
    return _gas_post({"action": "set_license", "value": value}) is not None

def gas_get_history():
    d = _gas_post({"action": "get_history"})
    rows = (d or {}).get("rows", [])
    return pd.DataFrame(rows, columns=HIST_HEADERS) if rows else pd.DataFrame(columns=HIST_HEADERS)

def gas_upsert_history(new_row):
    d = _gas_post({"action": "upsert_history", "row": new_row})
    if not d:
        return None
    rows = d.get("rows", [])
    return pd.DataFrame(rows, columns=HIST_HEADERS) if rows else pd.DataFrame(columns=HIST_HEADERS)

@st.cache_data(show_spinner=False)
def gas_get_template_bytes():
    d = _gas_post({"action": "get_template"})
    if not d or not d.get("b64"):
        return None
    try:
        return base64.b64decode(d["b64"])
    except Exception:
        return None

def gas_ping():
    """Return a human-readable diagnosis of the web-app connection."""
    url, token = _gas_cfg()
    if not url or not token:
        return "No URL/token found in Streamlit secrets (need [gas] url and token)."
    try:
        body = _json.dumps({"action": "get_license", "token": token})
        r = requests.post(url, data=body, headers={"Content-Type": "application/json"}, timeout=20)
        try:
            d = r.json()
            if d.get("ok"):
                return f"HTTP {r.status_code} — OK. Connection works."
            return f"HTTP {r.status_code} — reached the script but it returned error='{d.get('error')}' (usually a token mismatch)."
        except Exception:
            snippet = " ".join(r.text.split())[:180]
            return (f"HTTP {r.status_code} — response was not JSON. Usually means the deployment's access "
                    f"isn't set to 'Anyone', or the URL isn't the /exec deployment URL. Begins: {snippet}")
    except Exception as e:
        return f"Request failed: {e}"

# ── Parsing / multipliers ───────────────────────────────────────────────────────
def get_multiplier(text):
    s = str(text).lower().replace(" ", "")
    if "1/2bbl" in s: return 0.5
    if "1/6bbl" in s: return 0.166667
    if "can" in str(text).lower(): return 0.07258
    return None

def multiplier_for(product, memo):
    m = get_multiplier(product)
    return m if m is not None else get_multiplier(memo)

def norm(name):
    return str(name).lower().split(":")[0].strip()

def category_of(product):
    return str(product).split(":")[0].strip()

def classify_category(cat):
    c = str(cat).strip().lower()
    if c in ("", "nan") or "kegs" in c or "tap handle" in c or "service" in c:
        return "exclude"
    if "self distribute" in c or c.startswith("2"):
        return "self"
    if "brewpub" in c or "taproom" in c:
        return "brewpub"
    if "wv distributor" in c or c.startswith("1"):
        return "wv_dist"
    if "distributor" in c or c.startswith("3") or c.startswith("4"):
        return "other_state"
    return "unclassified"

CHANNEL_LABEL = {
    "wv_dist": "WV distributors (Q8/Q9)", "self": "Self-distribution (Q6/Q7)",
    "brewpub": "Brewpub (Q10/Q11)", "other_state": "Other state (total only)",
    "exclude": "Excluded (deposits/merch/services)", "unclassified": "UNCLASSIFIED - please review",
}

def _is_csv(f):
    return str(getattr(f, "name", "")).lower().endswith(".csv")

def _read_grid(uploaded_file, skiprows=0):
    uploaded_file.seek(0)
    if _is_csv(uploaded_file):
        return pd.read_csv(uploaded_file, header=None, skiprows=skiprows, dtype=object)
    return pd.read_excel(uploaded_file, header=None, skiprows=skiprows)

def detect_header_row(uploaded_file):
    raw = _read_grid(uploaded_file)
    for i, row in raw.iterrows():
        vals = [str(v).lower() for v in row.values]
        if any("transaction date" in v or "quantity" in v for v in vals):
            return i
    return 4

def load_template_bytes(override_file=None):
    if override_file is not None:
        return override_file.getvalue()
    if TEMPLATE_PATH and TEMPLATE_PATH.exists():
        return TEMPLATE_PATH.read_bytes()
    return None

def parse_sales_file(uploaded_file):
    uploaded_file.seek(0)
    header_row = detect_header_row(uploaded_file)
    raw_data = _read_grid(uploaded_file, skiprows=header_row + 1)
    num_cols = raw_data.shape[1]

    if num_cols >= 10:
        raw_data.columns = ["Customer", "Trans_Date", "Trans_Type", "Num", "Product",
                            "Memo", "Quantity", "Sales_Price", "Amount", "Balance"]
        product_col = "Product"
        customers, last_top, last_customer = [], None, None
        for _, row in raw_data.iterrows():
            val = row["Customer"]
            if pd.notna(val) and pd.isna(row["Trans_Date"]):
                s = str(val)
                if not s.lower().startswith("total"):
                    n = norm(s)
                    if n in st.session_state.abca_map:
                        last_top = s
                        last_customer = s
                    else:
                        same_stem = (last_top and s.strip().lower().split()[:1]
                                     == str(last_top).strip().lower().split()[:1])
                        last_customer = last_top if same_stem else s
            customers.append(last_customer)
        raw_data["Customer_filled"] = customers
    else:
        raw_data.columns = ["Customer", "Trans_Date", "Num", "ABCA", "Customer_orig", "Memo", "Quantity"]
        product_col = "Memo"
        raw_data["Product"] = raw_data["Memo"]
        raw_data["Customer_filled"] = raw_data["Customer_orig"].where(raw_data["Customer_orig"].notna()).ffill()

    df = raw_data.copy()
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce")
    df["Trans_Date"] = pd.to_datetime(df["Trans_Date"], errors="coerce")
    df = df.dropna(subset=["Trans_Date", "Quantity"])
    df = df[df["Quantity"] != 0]
    df["Quantity"] = df["Quantity"].astype(int)
    df["Num"] = df["Num"].astype(str).str.strip()
    df["Multiplier"] = df.apply(lambda r: multiplier_for(r[product_col], r.get("Memo")), axis=1)
    df["BBL"] = (df["Quantity"] * df["Multiplier"]).round(4)
    df["Gallons"] = (df["BBL"] * GAL_PER_BBL).round(2)
    df["Category"] = df[product_col].apply(category_of)
    df["Channel"] = df["Category"].apply(classify_category)
    df["norm"] = df["Customer_filled"].apply(norm)
    df["ABCA_Final"] = df["norm"].map(lambda x: st.session_state.abca_map.get(x, (None, None))[0])
    df["Licensee"] = df["norm"].map(lambda x: st.session_state.abca_map.get(x, (None, None))[1])
    return df, product_col

def _net_rows(df):
    """Keep every invoice line listed. For a distributor with returns (negative
    lines), remove the negative lines and subtract their total from that
    distributor's positive invoices (largest first, so no line goes negative) —
    the invoice list stays intact, just with adjusted barrels. net_negative lists
    distributors whose returns exceed all their sales for the month."""
    rows, net_negative = [], []
    d = df.copy()
    for r in d[d["ABCA_Final"].isna()].itertuples(index=False):   # unmapped pass-through
        rows.append([r.Trans_Date, r.Num, r.ABCA_Final, r.Licensee, round(r.BBL, 4)])
    mapped = d[d["ABCA_Final"].notna()]
    for _, g in mapped.groupby("ABCA_Final", sort=False):
        neg_total = g.loc[g["BBL"] < 0, "BBL"].sum()   # <= 0
        if neg_total >= 0:
            for r in g.itertuples(index=False):        # no returns → keep every line as-is
                rows.append([r.Trans_Date, r.Num, r.ABCA_Final, r.Licensee, round(r.BBL, 4)])
            continue
        deficit = -neg_total                           # amount to remove from positives
        pos = g[g["BBL"] > 0]
        adj = {idx: float(v) for idx, v in pos["BBL"].items()}
        for idx in pos.sort_values("BBL", ascending=False).index:   # largest first
            if deficit <= 1e-9:
                break
            take = min(adj[idx], deficit)
            adj[idx] = round(adj[idx] - take, 4)
            deficit = round(deficit - take, 4)
        for r in pos.sort_values("Trans_Date").itertuples():        # keep positive lines, listed
            rows.append([r.Trans_Date, r.Num, r.ABCA_Final, r.Licensee, round(adj[r.Index], 4)])
        if deficit > 1e-9:                             # returns exceeded all sales → unavoidable negative
            last = g.sort_values("Trans_Date").iloc[-1]
            rows.append([last["Trans_Date"], last["Num"], last["ABCA_Final"], last["Licensee"], round(-deficit, 4)])
            net_negative.append(str(g["Licensee"].iloc[0]))
    rows.sort(key=lambda x: (pd.isna(x[0]), x[0]))     # chronological
    return rows, net_negative

def process_file(uploaded_file, template_bytes):
    df, product_col = parse_sales_file(uploaded_file)
    # One-month guard: a monthly file must contain a single calendar month.
    months = sorted(df["Trans_Date"].dt.to_period("M").dropna().unique())
    if not months:
        raise ValueError("No dated transactions found in this file.")
    if len(months) > 1:
        names = ", ".join(p.to_timestamp().strftime("%b %Y") for p in months)
        raise ValueError(f"File spans multiple months ({names}). Upload one month per file.")
    month_label = months[0].to_timestamp().strftime("%b %Y")   # e.g. "Jun 2026"
    df = df[df["Channel"] != "exclude"]   # keg deposits / tap handles / services aren't taxable beer
    unmapped = [c for c in df[df["ABCA_Final"].isna()]["Customer_filled"].unique().tolist()
                if pd.notna(c) and str(c).strip().lower() != "nan"]
    rows, net_negative = _net_rows(df)
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb["Section_2"]
    for i, (dt, num, abca, lic, bbl) in enumerate(rows, 6):
        ws.cell(row=i, column=1, value=dt.date() if pd.notna(dt) else None)
        ws.cell(row=i, column=1).number_format = "MM/DD/YYYY"
        ws.cell(row=i, column=2, value=num)
        ws.cell(row=i, column=3, value=abca)
        ws.cell(row=i, column=4, value=lic)
        ws.cell(row=i, column=5, value=bbl)
        ws.cell(row=i, column=5).number_format = "0.0000"
    out = io.BytesIO(); wb.save(out); out.seek(0)
    total_bbl = round(sum(b for *_, b in rows if pd.notna(b)), 4)
    return out.read(), len(rows), total_bbl, unmapped, month_label, net_negative

def make_zip(results):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            if not r.get("error") and not r.get("unmapped"):
                zf.writestr(r["output_name"], r["result_bytes"])
    buf.seek(0)
    return buf.read()

# ── EOP helpers ──────────────────────────────────────────────────────────────────
def summarize_year(df):
    def gb(d):
        return round(d["Gallons"].sum(), 2), round(d["BBL"].sum(), 4)
    by_channel = {ch: gb(df[df["Channel"] == ch]) for ch in CHANNEL_LABEL}
    cat = (df.assign(Channel_label=df["Channel"].map(CHANNEL_LABEL))
             .groupby(["Category", "Channel_label"], as_index=False)
             .agg(Gallons=("Gallons", "sum"), BBL=("BBL", "sum"))
             .sort_values("Gallons", ascending=False, key=lambda s: s.abs())
             .round({"Gallons": 2, "BBL": 4}))
    beer = df[df["Channel"].isin(["wv_dist", "self", "brewpub", "other_state"])]
    total_g, total_b = gb(beer)
    return {"by_channel": by_channel, "cat_table": cat, "total_g": total_g, "total_b": total_b,
            "period_start": df["Trans_Date"].min(), "period_end": df["Trans_Date"].max(),
            "row_count": len(df), "has_unclassified": bool((df["Channel"] == "unclassified").any())}

def build_eop_summary(v):
    wb = Workbook(); ws = wb.active; ws.title = "EOP_Summary"
    rows = [
        ["WV Estimate/Report of Production", ""],
        ["Brewer", v.get("brewer", BREWER_NAME)],
        ["License Number", v.get("license", LICENSE_NUMBER)],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Fiscal Year", v["fiscal_year"]],
        ["Data period (from file)", f"{v['period_start']:%m/%d/%Y} - {v['period_end']:%m/%d/%Y}"],
        ["", ""],
        ["Form line", "Answer"],
        ["1. Est. gallons produced this year", v["q1_gal"]],
        ["2. Est. barrels produced this year (gal / 31)", v["q2_bbl"]],
        ["3. Production capacity (barrels)", v["q3_bbl"]],
        ["   Production capacity (gallons)", v["q3_bbl"] * GAL_PER_BBL],
        ["4. Prior year total production (gallons)", v["q4_gal"]],
        ["5. Prior year total SALES volume (gallons)", v["total_g"]],
        ["   Prior year total sales volume (barrels)", v["total_b"]],
        ["6. Self-distributed (gallons)", v["self_g"]],
        ["7. Self-distributed (barrels)", v["self_b"]],
        ["8. Sold to WV distributors (gallons)", v["dist_g"]],
        ["9. Sold to WV distributors (barrels)", v["dist_b"]],
        ["10. Sold through brewpub (gallons)", v["bp_g"]],
        ["11. Sold through brewpub (barrels)", v["bp_b"]],
        ["", ""],
        ["(Out-of-state sales, in total but no WV line) gallons", v["other_g"]],
    ]
    for r in rows: ws.append(r)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 28
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()

# ── Connect + resolve license ────────────────────────────────────────────────────
sheets_ok = storage_ready()
if "license_no" not in st.session_state:
    st.session_state.license_no = (gas_get_license() or LICENSE_NUMBER) if sheets_ok else LICENSE_NUMBER

# ── UI ────────────────────────────────────────────────────────────────────────────
st.title("🍺 WV BBL Tax Reporter")
st.markdown(f"*{BREWER_NAME} — License #{st.session_state.license_no}*")
st.markdown("---")

tab_tax, tab_eop = st.tabs(["Monthly BBL Tax Report", "Production Estimate/Report"])

# ════════════════════════════════════════════════════════════════════════════════
#  TAB 1 — Monthly BBL Tax Report
# ════════════════════════════════════════════════════════════════════════════════
with tab_tax:
    st.subheader("1. WV Upload Template")
    repo_ok = bool(TEMPLATE_PATH and TEMPLATE_PATH.exists())
    if repo_ok:
        source, label = TEMPLATE_PATH.read_bytes(), f"repo — {TEMPLATE_PATH.name}"
    elif EMBEDDED_TEMPLATE:
        source, label = EMBEDDED_TEMPLATE, "built into the app"
    else:
        source, label = None, None
    if source is not None:
        st.markdown(f'<div class="template-saved">✓ Template loaded ({label}) — '
                    'nothing to upload.</div>', unsafe_allow_html=True)
        override = st.file_uploader("Override template for this session only (optional)",
                                    type=["xlsx"], key="template_override")
    else:
        st.markdown('<div class="error-box">⚠️ No template available yet — upload one below for this '
                    'session.</div>', unsafe_allow_html=True)
        override = st.file_uploader("Upload __WV_Upload_Template.xlsx", type=["xlsx"], key="template_override")
    template_bytes = override.getvalue() if override else source

    st.markdown("---")
    st.subheader("2. Upload Monthly Sales File(s)")
    sales_files = st.file_uploader("Upload one or more monthly sales exports (.xlsx or .csv)",
                                   type=["xlsx", "csv"], accept_multiple_files=True, key="sales")

    st.markdown("---")
    with st.expander("⚙️ Manage Distributor ABCA Numbers"):
        abca_df = pd.DataFrame([{"Distributor (lowercase key)": k, "ABCA License #": val[0], "Display Name": val[1]}
                                for k, val in st.session_state.abca_map.items()])
        st.dataframe(abca_df, use_container_width=True, hide_index=True)
        col1, col2, col3 = st.columns(3)
        with col1: new_key = st.text_input("Customer name (as in QuickBooks)", key="new_key")
        with col2: new_abca = st.text_input("ABCA License #", key="new_abca")
        with col3: new_display = st.text_input("Display Name", key="new_display")
        if st.button("Add Distributor"):
            if new_key and new_abca and new_display:
                st.session_state.abca_map[norm(new_key)] = (new_abca, new_display)
                st.success(f"Added: {new_display}"); st.rerun()
            else:
                st.warning("Please fill in all three fields.")

    st.markdown("---")
    st.subheader("3. Generate Reports")
    if st.button("▶ Process Files", disabled=(not template_bytes or not sales_files)):
        st.session_state.results = []
        for sf in sales_files:
            base = sf.name.rsplit(".", 1)[0]
            try:
                sf.seek(0)
                rb, rc, tb, um, mlabel, netneg = process_file(sf, template_bytes)
                output_name = f"{mlabel} WV BBL Tax.xlsx"
                st.session_state.results.append({"filename": base, "output_name": output_name,
                    "result_bytes": rb, "row_count": rc, "total_bbl": tb, "unmapped": um,
                    "net_negative": netneg, "error": None})
            except Exception as e:
                st.session_state.results.append({"filename": base, "output_name": f"{base}_Final.xlsx",
                    "unmapped": [], "error": str(e)})

    if st.session_state.results:
        successful = [r for r in st.session_state.results if not r.get("error") and not r.get("unmapped")]
        if len(successful) > 1:
            st.download_button(f"⬇ Download All {len(successful)} Files as ZIP", data=make_zip(successful),
                file_name="BBL_Reports.zip", mime="application/zip", key="dl_zip")
            st.markdown("&nbsp;")
        for r in st.session_state.results:
            if r.get("error"):
                st.markdown(f'<div class="error-box">✗ <b>{r["filename"]}</b> — Error: {r["error"]}</div>', unsafe_allow_html=True)
            elif r.get("unmapped"):
                u = r["unmapped"]
                st.markdown(f'<div class="error-box">⚠️ <b>{r["filename"]}</b> — {len(u)} unmapped: {", ".join(str(x) for x in u)}<br>Add them in the ABCA manager above and re-run.</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="result-box">✓ <b>{r["output_name"]}</b> — {r["row_count"]} rows &nbsp;|&nbsp; {r["total_bbl"]} BBL total</div>', unsafe_allow_html=True)
                if r.get("net_negative"):
                    st.markdown(f'<div class="error-box">⚠️ Net-negative this month: {", ".join(r["net_negative"])}. '
                                'The state can\'t accept a negative line — carry that credit to another month or adjust it manually before uploading.</div>',
                                unsafe_allow_html=True)
                st.download_button(f"⬇ Download {r['output_name']}", data=r["result_bytes"], file_name=r["output_name"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_{r['filename']}")
    elif not template_bytes or not sales_files:
        st.caption("Upload a template and at least one sales file to enable processing.")

# ════════════════════════════════════════════════════════════════════════════════
#  TAB 2 — WV Estimate/Report of Production
# ════════════════════════════════════════════════════════════════════════════════
with tab_eop:
    st.subheader("WV Estimate/Report of Production")
    st.markdown(f'<div class="result-box">Brewer: <b>{BREWER_NAME}</b></div>', unsafe_allow_html=True)

    license_no = st.text_input("WV brewer license #", value=st.session_state.license_no, key="license_input")
    if license_no != st.session_state.license_no:
        st.session_state.license_no = license_no
        if sheets_ok and gas_set_license(license_no):
            st.toast("License number saved to the sheet")

    if sheets_ok:
        st.markdown('<div class="template-saved">✓ Sheet credentials loaded — use "Test connection" below '
                    'to confirm read/write actually works.</div>', unsafe_allow_html=True)
        with st.expander("Test connection"):
            if st.button("Run connection test", key="gas_ping_btn"):
                st.code(gas_ping())
    else:
        st.markdown('<div class="error-box">Sheet not connected yet — using the built-in license default '
                    'and the file fallback for history. Add the web-app URL + token to Streamlit secrets to '
                    'turn on auto-save.</div>', unsafe_allow_html=True)

    st.caption("Channel is read from the product category: 1 = WV distributors, 2 = self-distribute, "
               "3/4 = other states. Keg deposits (6), tap handles (7) and services are ignored. "
               "Barrels are 31 gallons.")

    fy = st.text_input("Fiscal year label", value="", placeholder="e.g. 2025-2026", key="fy")
    year_file = st.file_uploader("Upload the year's sales export (raw dump .xlsx or .csv)", type=["xlsx", "csv"], key="year_file")
    prior_history = None
    if not sheets_ok:
        prior_history = st.file_uploader(
            "Optional: prior production_history.xlsx (fallback when the sheet is offline)", type=["xlsx"], key="hist")

    if st.button("▶ Analyze Year", disabled=(not year_file)):
        try:
            year_file.seek(0)
            df, _ = parse_sales_file(year_file)
            summary = summarize_year(df)
            summary["prior_history_bytes"] = prior_history.getvalue() if prior_history else None
            st.session_state.eop = summary
        except Exception as e:
            st.session_state.eop = None
            st.markdown(f'<div class="error-box">✗ Error: {e}</div>', unsafe_allow_html=True)

    v = st.session_state.eop
    if v:
        bc = v["by_channel"]
        wv_g, wv_b = bc["wv_dist"]; self_g, self_b = bc["self"]
        bp_data_g, bp_data_b = bc["brewpub"]; other_g, other_b = bc["other_state"]

        st.markdown(f'<div class="result-box">Period in file: '
                    f'<b>{v["period_start"]:%m/%d/%Y}</b> – <b>{v["period_end"]:%m/%d/%Y}</b> '
                    f'&nbsp;|&nbsp; {v["row_count"]} line items</div>', unsafe_allow_html=True)

        if v["has_unclassified"]:
            st.markdown('<div class="error-box">⚠️ One or more product categories couldn\'t be classified '
                        'and were left OUT of the totals. See the category audit below.</div>', unsafe_allow_html=True)
        with st.expander("Category audit (every category and where it landed)", expanded=v["has_unclassified"]):
            st.dataframe(v["cat_table"], use_container_width=True, hide_index=True)

        # Brewpub (Q10/Q11): entered in BARRELS, converts to gallons, rolls into the total
        bp_b_final = st.number_input("Brewpub sales (barrels) — Q10/Q11, manual entry",
            min_value=0.0, value=float(bp_data_b), step=1.0, key="bp_bbl",
            help="No brewpub category exists in the sales data, so enter it here. It converts to gallons "
                 "and is added to the totals below.")
        bp_g_final = round(bp_b_final * GAL_PER_BBL, 2)
        st.caption(f"= {bp_g_final:,.2f} gallons — added to the total below.")
        total_b_final = round(v["total_b"] - bp_data_b + bp_b_final, 4)
        total_g_final = round(v["total_g"] - bp_data_g + bp_g_final, 2)

        st.markdown("**Channel breakdown:**")
        breakdown = pd.DataFrame([
            {"Channel": "Sold to WV distributors", "Gallons": wv_g, "Barrels": wv_b},
            {"Channel": "Self-distributed",        "Gallons": self_g, "Barrels": self_b},
            {"Channel": "Brewpub (manual)",        "Gallons": bp_g_final, "Barrels": bp_b_final},
            {"Channel": "Other state (total only)","Gallons": other_g, "Barrels": other_b},
            {"Channel": "TOTAL beer sales",        "Gallons": total_g_final, "Barrels": total_b_final},
        ])
        st.dataframe(breakdown, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("**Production-side figures:**")
        cga, cgb = st.columns(2)
        with cga:
            q1_override = st.number_input("Q1 — Est. gallons produced this year (0 = use total sales)",
                min_value=0.0, value=0.0, step=100.0, key="q1_override",
                help="Leave at 0 to use total sales (incl. brewpub), which updates live as you change "
                     "the brewpub barrels. Type a number to override with your own estimate.")
        with cgb:
            q4_gal = st.number_input("Q4 — Prior year total production (gallons)",
                min_value=0.0, value=0.0, step=100.0, key="q4")
        q1_gal = q1_override if q1_override > 0 else total_g_final
        q2_bbl = round(q1_gal / GAL_PER_BBL, 2)
        q1_note = "" if q1_override > 0 else " (= total sales, auto)"
        st.markdown(f'<div class="result-box">'
                    f'Q1 — Est. gallons produced: <b>{q1_gal:,.2f}</b> gal{q1_note}<br>'
                    f'Q2 — Est. barrels produced (Q1 ÷ 31): <b>{q2_bbl}</b> bbl<br>'
                    f'Q3 — Production capacity (constant): <b>{FACILITY_CAPACITY_BBL:,}</b> bbl '
                    f'({FACILITY_CAPACITY_BBL * GAL_PER_BBL:,} gal)</div>', unsafe_allow_html=True)

        fiscal_year = fy or f"{v['period_start']:%Y}-{v['period_end']:%Y}"
        values = {
            "fiscal_year": fiscal_year, "period_start": v["period_start"], "period_end": v["period_end"],
            "brewer": BREWER_NAME, "license": st.session_state.license_no,
            "q1_gal": q1_gal, "q2_bbl": q2_bbl, "q3_bbl": FACILITY_CAPACITY_BBL, "q4_gal": q4_gal,
            "total_g": total_g_final, "total_b": total_b_final,
            "self_g": self_g, "self_b": self_b, "dist_g": wv_g, "dist_b": wv_b,
            "bp_g": bp_g_final, "bp_b": bp_b_final, "other_g": other_g, "other_b": other_b,
        }

        st.markdown("---")
        st.download_button("⬇ Download EOP Summary (all 11 lines, labeled)", data=build_eop_summary(values),
            file_name=f"WV_Production_{fiscal_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_eop")

        new_row = {
            "Fiscal Year": fiscal_year,
            "Total Sales (gallons)": float(total_g_final), "Total Sales (barrels)": float(total_b_final),
            "WV Distributor (gallons)": float(wv_g), "Self-distributed (gallons)": float(self_g),
            "Brewpub (gallons)": float(bp_g_final), "Other State (gallons)": float(other_g),
            "Generated": datetime.now().strftime("%Y-%m-%d"),
        }

        st.markdown("---")
        st.markdown("**Year-to-year history (gallons):**")
        if sheets_ok:
            if st.button(f"💾 Save / update {fiscal_year} in the sheet"):
                updated = gas_upsert_history(new_row)
                if updated is not None:
                    st.toast(f"{fiscal_year} saved to the sheet")
                else:
                    st.markdown('<div class="error-box">Could not write to the sheet — check the web-app URL/token.</div>',
                                unsafe_allow_html=True)
            st.dataframe(gas_get_history(), use_container_width=True, hide_index=True)
            st.caption("History lives in 'MSBC WV Production Data' on your Drive — no files to download.")
        else:
            if v.get("prior_history_bytes"):
                try:
                    hist = pd.read_excel(io.BytesIO(v["prior_history_bytes"]))
                except Exception:
                    hist = pd.DataFrame(columns=HIST_HEADERS)
            else:
                hist = pd.DataFrame(columns=HIST_HEADERS)
            if "Fiscal Year" in hist.columns:
                hist = hist[hist["Fiscal Year"].astype(str) != str(fiscal_year)]
            hist = pd.concat([hist, pd.DataFrame([new_row])], ignore_index=True).sort_values("Fiscal Year").reset_index(drop=True)
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as w:
                hist.to_excel(w, index=False, sheet_name="Production_History")
            out.seek(0)
            st.dataframe(hist, use_container_width=True, hide_index=True)
            st.download_button("⬇ Download updated production_history.xlsx", data=out.read(),
                file_name="production_history.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_hist")
            st.caption("Fallback mode. Connect the web app to drop the file step.")
    elif not year_file:
        st.caption("Upload the year's raw-dump export to enable analysis.")
